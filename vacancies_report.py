import csv
import os
import re
import time
from itertools import islice

import numpy as np
import openpyxl
from matplotlib import pyplot as plt
from openpyxl.styles import Side, Font, Border, Alignment
from openpyxl.utils import get_column_letter

currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                   "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}

fileName = input("Введите название файла: ")
input_vacancy_name = input("Введите название профессии: ")
output_view = input("Вывести данные в виде:(Таблица/Графики) ")


class Vacancy(object):
    def __init__(self, vacancy_list):
        self.name = vacancy_list[0]
        self.salary_from = float(vacancy_list[1])
        self.salary_to = float(vacancy_list[2])
        self.salary_currency = vacancy_list[3]
        self.area_name = vacancy_list[4]
        self.published_at = int(vacancy_list[5][0:4])

    def get_list(self):
        return [self.name, self.salary_from, self.salary_to, self.salary_currency, self.area_name, self.published_at]


class DataSet(object):
    def __init__(self, filename):
        self.file_name = filename
        self.vacancies_objects = []
        DataSet.csv1(self)

    def csv1(self):
        file_csv = open(self.file_name, 'r', encoding="utf-8-sig")
        reader_csv = list(csv.reader(file_csv))
        for vacancy in reader_csv[1:]:
            if "" not in vacancy and len(vacancy) == 6:
                for i in range(len(vacancy)):
                    if vacancy[i].__contains__('\n'):
                        vacancy[i] = '!'.join(vacancy[i].split('\n'))
                    else:
                        vacancy[i] = " ".join(re.sub(r'\<[^>]*\>', '', vacancy[i]).split())
                if int(vacancy[5][0:4]) < 2015:
                    self.vacancies_objects.append(Vacancy(vacancy))


class InputConnect(object):
    def __init__(self, vacancy_name):
        self.dict1, self.dict2, self.dict3, self.dict4, self.dict5, self.dict6 = None, None, None, None, None, None
        self.vacancy_name = vacancy_name.lower()
        self.data = DataSet(fileName)
        self.year_stat = {}
        self.vacancy_stat = {}
        self.cities_salary = {}
        self.count = len(self.data.vacancies_objects)

    def filtration(self):
        self.data.vacancies_objects = [x for x in self.data.vacancies_objects if
                                       self.vacancy_name in x.name.lower()]

    def level_year_stat(self, dictionary):
        for vac in self.data.vacancies_objects:
            if vac.published_at not in dictionary:
                dictionary[vac.published_at] = [
                    (vac.salary_from + vac.salary_to) / 2 * currency_to_rub[vac.salary_currency], 1]

            else:
                dictionary[vac.published_at][0] += (vac.salary_from + vac.salary_to) / 2 * currency_to_rub[
                    vac.salary_currency]
                dictionary[vac.published_at][1] += 1

    def cities_sal(self):
        for vac in self.data.vacancies_objects:
            if vac.area_name not in self.cities_salary:
                self.cities_salary[vac.area_name] = [
                    (vac.salary_from + vac.salary_to) / 2 * currency_to_rub[vac.salary_currency], 1]
            else:
                self.cities_salary[vac.area_name][0] += (vac.salary_from + vac.salary_to) / 2 * currency_to_rub[
                    vac.salary_currency]
                self.cities_salary[vac.area_name][1] += 1

    def create_output_statistics(self):
        self.dict1 = {a: round(self.year_stat[a][0] / self.year_stat[a][1]) for a in self.year_stat}
        self.dict2 = {a: self.year_stat[a][1] for a in self.year_stat}
        self.dict3 = {a: round(self.vacancy_stat[a][0] / self.vacancy_stat[a][1]) for a in self.vacancy_stat}
        self.dict4 = {a: self.vacancy_stat[a][1] for a in self.vacancy_stat}
        self.dict5 = dict(islice(sorted(
            {a: round(self.cities_salary[a][0] / self.cities_salary[a][1]) for a in self.cities_salary if
             self.cities_salary[a][1] / self.count > 0.01}.items(),
            key=lambda item: item[1],
            reverse=True), 10))
        self.dict6 = dict(islice(sorted(
            {a: round(self.cities_salary[a][1] / self.count, 4) for a in self.cities_salary if
             self.cities_salary[a][1] / self.count > 0.01}.items(),
            key=lambda item: item[1],
            reverse=True), 10))

    def print_statistics(self):
        print('Динамика уровня зарплат по годам: ', self.dict1)
        print('Динамика количества вакансий по годам: ', self.dict2)
        print('Динамика уровня зарплат по годам для выбранной профессии: ', self.dict3)
        print('Динамика количества вакансий по годам для выбранной профессии: ', self.dict4)
        print('Уровень зарплат по городам (в порядке убывания): ', self.dict5)
        print('Доля вакансий по городам (в порядке убывания): ', self.dict6)

    def get_statistics(self):
        return [[self.dict1, self.dict3, self.dict2, self.dict4], [self.dict5, self.dict6]]


class Report(object):
    def __init__(self, dictionary_list):
        self.statistics_by_countries = dictionary_list[1]
        self.statistics_by_years = dictionary_list[0]

    def generate_excel(self):
        self.wb = openpyxl.Workbook()
        self.sheet1 = self.wb.active
        self.sheet1.title = "Статистика по годам"
        self.sheet1.append(('Год', 'Средняя зарплата', f'Средняя зарплата - {input_vacancy_name}', 'Количество вакансий',
                            f'Количество вакансий - {input_vacancy_name}'))
        self.sheet2 = self.wb.create_sheet("Статистика по городам")
        self.sheet2.append(('Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'))
        Report.table_fill(self.sheet1, 0, self.statistics_by_years[0])
        for i in range(1, 5):
            Report.table_fill(self.sheet1, i, self.statistics_by_years[i - 1])
        for i in range(0, 2):
            Report.table_fill(self.sheet2, i, self.statistics_by_countries[0])
        for i in range(3, 5):
            Report.table_fill(self.sheet2, i, self.statistics_by_countries[1])
        Report.sheet_format(self.sheet1)
        Report.sheet_format(self.sheet2)

        self.wb.save('report.xlsx')

    @staticmethod
    def table_fill(sheets, index, dictionary):
        a = 2
        for key, value in dictionary.items():
            if sheets[1][index].value == 'Год' or sheets[1][index].value == 'Город':
                sheets[a][index].value = key
            else:
                sheets[a][index].value = "{:.2%}".format(value) if value < 1 else value
            a += 1

    @staticmethod
    def sheet_format(sheets):
        thins = Side(border_style="thin", color="000000")
        for column in sheets.columns:
            max_length = 0
            for cell in column:
                if cell.row == 1:
                    cell.font = Font(bold=True)
                if cell.value:
                    max_length = len(str(cell.value)) + 2 if len(str(cell.value)) + 2 > max_length else max_length
                    cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)
                else:
                    max_length = 2
                if cell.column > 1 and cell.row > 1 and column[0].value != 'Город':
                    cell.alignment = Alignment(horizontal='right')
            sheets.column_dimensions[get_column_letter(cell.column)].width = max_length


class GraphReport(object):
    def __init__(self, dictionary_list):
        self.statistics_by_countries = dictionary_list[1]
        self.statistics_by_years = dictionary_list[0]

    def generate_graph(self):
        self.y = np.arange(len(self.statistics_by_countries[0].keys()))
        self.x = np.arange(len(self.statistics_by_years[0].keys()))
        self.width = 0.4
        self.fig, self.axs = plt.subplots(nrows=2, ncols=2, figsize=(8, 6))
        GraphReport.salary_level_create(self)
        GraphReport.vacancy_count_create(self)
        GraphReport.countries_salary_create(self)
        GraphReport.part_countries_create(self)
        self.fig.tight_layout()
        plt.show()


    def salary_level_create(self):
        self.axs[0, 0].bar(self.x - self.width / 2, self.statistics_by_years[0].values(), self.width, label='средняя з/п')
        self.axs[0, 0].bar(self.x + self.width / 2, self.statistics_by_years[1].values(), self.width, label='з/п программист')
        self.axs[0, 0].set_title('Уровень зарплат по годам', fontsize=14)
        self.axs[0, 0].set_xticks(self.x, self.statistics_by_years[0].keys(), rotation=90, fontsize=8)
        self.axs[0, 0].grid(axis='y')
        self.axs[0, 0].legend(fontsize=8)

    def vacancy_count_create(self):
        self.axs[0, 1].bar(self.x - self.width / 2, self.statistics_by_years[2].values(), self.width, label='количество вакансий')
        self.axs[0, 1].bar(self.x + self.width / 2, self.statistics_by_years[3].values(), self.width, label='количество вакансий программист')
        self.axs[0, 1].set_title('Количество вакансий по годам', fontsize=14)
        self.axs[0, 1].set_xticks(self.x, self.statistics_by_years[0].keys(), rotation=90, fontsize=8)
        self.axs[0, 1].grid(axis='y')
        self.axs[0, 1].legend(fontsize=8, loc='upper left')

    def countries_salary_create(self):
        self.axs[1, 0].barh(self.y - self.width / 2, self.statistics_by_countries[0].values(), self.width * 2)
        self.axs[1, 0].set_title('Уровень зарплат по городам', fontsize=14)
        self.axs[1, 0].set_yticks(self.y, self.statistics_by_countries[0].keys(), fontsize=8)
        self.axs[1, 0].grid(axis='x')
        self.axs[1, 0].invert_yaxis()

    def part_countries_create(self):
        arg = [x * 100 for x in self.statistics_by_countries[1].values()]
        arg.append(100 - sum(arg))
        arg1 = list(self.statistics_by_countries[0].keys())
        arg1.append('Другие')
        self.axs[1, 1].pie(arg, labels=arg1, textprops={'fontsize': 6})
        self.axs[1, 1].set_title('Количество вакансий по годам', fontsize=14)


if os.path.getsize(fileName) != 0:
    print_tab = InputConnect(input_vacancy_name)
    print_tab.level_year_stat(print_tab.year_stat)
    print_tab.cities_sal()
    print_tab.filtration()
    print_tab.level_year_stat(print_tab.vacancy_stat)
    print_tab.create_output_statistics()
    print_tab.print_statistics()
    if output_view == 'Таблица':
        tab = Report(print_tab.get_statistics())
        tab.generate_excel()
    elif output_view == 'Графики':
        graph = GraphReport(print_tab.get_statistics())
        graph.generate_graph()
else:
    print('Пустой файл')

